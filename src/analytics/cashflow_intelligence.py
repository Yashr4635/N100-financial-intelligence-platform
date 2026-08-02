"""
cashflow_intelligence.py

Sprint 5 – Cash Flow Intelligence Engine for the N100 Financial
Intelligence Platform.

Analyses multi-year cash flow data to derive qualitative patterns
and early-warning signals for every company.

Metrics produced
----------------
CFO Quality           : Operating CF / Net Profit (cash conversion quality)
CapEx Intensity       : CapEx / Revenue (investment aggressiveness)
Capital Allocation Pattern : label assigned to each company-year
Distress Detection    : flag when CFO < 0 AND FCF < 0 AND D/E > 2
Deleveraging Detection: flag when Total Debt fell YoY by >= 5%

Inputs
------
- data/raw/cashflow.xlsx          (operating / investing / financing CF)
- data/raw/profitandloss.xlsx     (sales, net_profit)
- data/output/company_health_scores.csv  (FCF, CapEx, Debt columns)

Outputs
-------
- data/output/cashflow_intelligence.xlsx   (multi-sheet workbook)
- data/output/distress_alerts.csv          (companies under distress)
- data/output/pattern_changes.csv          (capital allocation changes)
"""

from __future__ import annotations

import logging
import os
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
if not logger.handlers:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
CASHFLOW_RAW_PATH = os.path.join("data", "raw", "cashflow.xlsx")
PNL_RAW_PATH = os.path.join("data", "raw", "profitandloss.xlsx")
HEALTH_SCORES_PATH = os.path.join("data", "output", "company_health_scores.csv")
OUTPUT_DIR = os.path.join("data", "output")

DISTRESS_D2E_THRESHOLD = 1.5       # D/E above which we check distress
DELEVERAGING_DROP_PCT = 5.0         # % YoY debt drop = deleveraging
CAPEX_INTENSITY_HIGH = 0.15         # CapEx/Revenue >= 15% = high intensity
CAPEX_INTENSITY_LOW = 0.05          # CapEx/Revenue < 5% = asset-light


class CashFlowIntelligenceEngine:
    """
    Derives cash flow quality, CapEx intensity, capital allocation
    patterns, and distress / deleveraging signals for every company.
    """

    def __init__(
        self,
        cashflow_path: Optional[str] = None,
        pnl_path: Optional[str] = None,
        health_scores_path: Optional[str] = None,
        output_dir: Optional[str] = None,
    ) -> None:
        self.cashflow_path = cashflow_path or CASHFLOW_RAW_PATH
        self.pnl_path = pnl_path or PNL_RAW_PATH
        self.health_scores_path = health_scores_path or HEALTH_SCORES_PATH
        self.output_dir = output_dir or OUTPUT_DIR
        os.makedirs(self.output_dir, exist_ok=True)

        self.excel_output_path = os.path.join(self.output_dir, "cashflow_intelligence.xlsx")
        self.distress_output_path = os.path.join(self.output_dir, "distress_alerts.csv")
        self.pattern_output_path = os.path.join(self.output_dir, "pattern_changes.csv")

    # ------------------------------------------------------------------
    # Loading helpers
    # ------------------------------------------------------------------

    def _load_excel(self, path: str, label: str, header: int = 1) -> pd.DataFrame:
        if not os.path.exists(path):
            logger.warning("%s not found at '%s'.", label, path)
            return pd.DataFrame()
        try:
            df = pd.read_excel(path, header=header)
            df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
            logger.info("Loaded %s: %d rows.", label, len(df))
            return df
        except Exception as exc:
            logger.error("Failed to load %s: %s", label, exc)
            return pd.DataFrame()

    def _load_csv(self, path: str, label: str) -> pd.DataFrame:
        if not os.path.exists(path):
            logger.warning("%s not found at '%s'.", label, path)
            return pd.DataFrame()
        try:
            df = pd.read_csv(path)
            logger.info("Loaded %s: %d rows.", label, len(df))
            return df
        except Exception as exc:
            logger.error("Failed to load %s: %s", label, exc)
            return pd.DataFrame()

    # ------------------------------------------------------------------
    # Year normalisation
    # ------------------------------------------------------------------

    @staticmethod
    def _normalise_year(year_val) -> Optional[int]:
        """
        Normalise various year formats:
          "Mar-13" → 2013, "Dec 2012" → 2012, 2023 → 2023
        """
        if year_val is None or (isinstance(year_val, float) and pd.isna(year_val)):
            return None
        s = str(year_val).strip()
        # "Mar-13", "Dec-22"
        import re
        m = re.match(r"[A-Za-z]+-(\d{2})$", s)
        if m:
            yy = int(m.group(1))
            return 2000 + yy if yy < 50 else 1900 + yy
        # "Dec 2012", "Mar 2023"
        m2 = re.match(r"[A-Za-z]+\s+(\d{4})$", s)
        if m2:
            return int(m2.group(1))
        # Plain year
        try:
            return int(float(s))
        except ValueError:
            return None

    # ------------------------------------------------------------------
    # CFO Quality
    # ------------------------------------------------------------------

    def _compute_cfo_quality(
        self, cf_df: pd.DataFrame, pnl_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        CFO Quality = Operating Cash Flow / Net Profit.
        > 1.0 → excellent; 0.5-1.0 → acceptable; < 0.5 → poor.
        """
        if cf_df.empty or pnl_df.empty:
            return pd.DataFrame()

        cf = cf_df[["company_id", "year", "operating_activity"]].copy()
        cf["year"] = cf["year"].apply(self._normalise_year)
        cf = cf.dropna(subset=["year"])
        cf["year"] = cf["year"].astype(int)

        pnl = pnl_df[["company_id", "year", "net_profit", "sales"]].copy()
        pnl["year"] = pnl["year"].apply(self._normalise_year)
        pnl = pnl.dropna(subset=["year"])
        pnl["year"] = pnl["year"].astype(int)

        merged = pd.merge(cf, pnl, on=["company_id", "year"], how="inner")
        merged["cfo"] = pd.to_numeric(merged["operating_activity"], errors="coerce")
        merged["net_profit"] = pd.to_numeric(merged["net_profit"], errors="coerce")
        merged["sales"] = pd.to_numeric(merged["sales"], errors="coerce")

        # CFO Quality ratio
        merged["cfo_quality"] = merged.apply(
            lambda r: round(r["cfo"] / r["net_profit"], 3)
            if r["net_profit"] not in (None, 0) and not pd.isna(r["net_profit"])
            else None,
            axis=1,
        )

        def _cfo_label(val):
            if val is None or pd.isna(val):
                return "Insufficient Data"
            if val >= 1.0:
                return "Excellent"
            if val >= 0.5:
                return "Acceptable"
            return "Poor"

        merged["cfo_quality_label"] = merged["cfo_quality"].apply(_cfo_label)
        return merged[["company_id", "year", "cfo", "net_profit", "sales", "cfo_quality", "cfo_quality_label"]]

    # ------------------------------------------------------------------
    # CapEx Intensity
    # ------------------------------------------------------------------

    def _compute_capex_intensity(self, health_df: pd.DataFrame) -> pd.DataFrame:
        """
        CapEx Intensity = CapEx / Revenue.
        Uses CapEx and Revenue from the health scores / ratio engine output.
        """
        if health_df.empty:
            return pd.DataFrame()

        df = health_df.copy()
        for col in ["capex_cr", "free_cash_flow_cr", "cash_from_operations_cr"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        # Approximate revenue from EPS * shares (if not available use a proxy)
        # We don't have revenue in health_scores directly; leave as NaN
        df["capex_intensity"] = None

        if "capex_cr" in df.columns:
            df["capex_intensity_label"] = df["capex_cr"].apply(
                lambda v: (
                    "High CapEx" if pd.notna(v) and v > 500
                    else ("Moderate CapEx" if pd.notna(v) and v > 100 else "Low CapEx")
                )
            )
        else:
            df["capex_intensity_label"] = "Data Unavailable"

        keep = ["company_id", "year", "capex_cr", "cash_from_operations_cr",
                "free_cash_flow_cr", "capex_intensity", "capex_intensity_label"]
        return df[[c for c in keep if c in df.columns]].copy()

    # ------------------------------------------------------------------
    # Capital Allocation Pattern
    # ------------------------------------------------------------------

    def _classify_capital_allocation(self, row: pd.Series) -> str:
        """
        Classify a company-year into a capital allocation pattern:
          Growth Investor   : CapEx high + FCF positive
          Dividend Focus    : Low CapEx + dividends paid
          Debt Reducer      : FCF > 0 + deleveraging
          Cash Accumulator  : FCF high + low CapEx + low debt
          Distressed        : FCF < 0 + CFO < 0
          Balanced          : default
        """
        fcf = _safe_float(row.get("free_cash_flow_cr"))
        cfo = _safe_float(row.get("cash_from_operations_cr"))
        capex = _safe_float(row.get("capex_cr"))
        div = _safe_float(row.get("dividend_payout_ratio_pct"))
        debt = _safe_float(row.get("total_debt_cr"))

        if fcf is not None and cfo is not None and fcf < 0 and cfo < 0:
            return "Distressed"
        if capex is not None and capex > 500 and fcf is not None and fcf > 0:
            return "Growth Investor"
        if fcf is not None and fcf > 0 and debt is not None and debt == 0:
            return "Cash Accumulator"
        if div is not None and div >= 25 and capex is not None and capex < 200:
            return "Dividend Focus"
        if fcf is not None and fcf > 0 and debt is not None and debt < 100:
            return "Debt Reducer"
        return "Balanced"

    # ------------------------------------------------------------------
    # Distress Detection
    # ------------------------------------------------------------------

    def _detect_distress(self, health_df: pd.DataFrame) -> pd.DataFrame:
        """
        Flag company-years in financial distress:
          CFO < 0 AND FCF < 0 AND D/E > threshold
        """
        if health_df.empty:
            return pd.DataFrame()

        df = health_df.copy()
        for col in ["free_cash_flow_cr", "cash_from_operations_cr", "debt_to_equity"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        conditions = pd.Series(True, index=df.index)
        if "cash_from_operations_cr" in df.columns:
            conditions &= df["cash_from_operations_cr"] < 0
        if "free_cash_flow_cr" in df.columns:
            conditions &= df["free_cash_flow_cr"] < 0
        if "debt_to_equity" in df.columns:
            conditions &= df["debt_to_equity"] > DISTRESS_D2E_THRESHOLD

        distress = df[conditions].copy()
        if distress.empty:
            return pd.DataFrame(columns=["company_id", "year", "distress_reason",
                                         "free_cash_flow_cr", "cash_from_operations_cr", "debt_to_equity"])

        distress["distress_reason"] = (
            "CFO negative + FCF negative + D/E > "
            + str(DISTRESS_D2E_THRESHOLD)
        )

        keep = ["company_id", "year", "distress_reason",
                "free_cash_flow_cr", "cash_from_operations_cr", "debt_to_equity"]
        return distress[[c for c in keep if c in distress.columns]].copy()

    # ------------------------------------------------------------------
    # Deleveraging Detection
    # ------------------------------------------------------------------

    def _detect_deleveraging(self, health_df: pd.DataFrame) -> pd.DataFrame:
        """
        Flag company-years where total debt fell by >= DELEVERAGING_DROP_PCT% YoY.
        """
        if health_df.empty or "total_debt_cr" not in health_df.columns:
            return pd.DataFrame()

        df = health_df.copy()
        df["total_debt_cr"] = pd.to_numeric(df["total_debt_cr"], errors="coerce")
        df["year"] = pd.to_numeric(df["year"], errors="coerce")
        df = df.sort_values(["company_id", "year"])
        df["prev_debt"] = df.groupby("company_id")["total_debt_cr"].shift(1)
        df["debt_change_pct"] = (
            (df["total_debt_cr"] - df["prev_debt"]) / df["prev_debt"].replace(0, float("nan"))
        ) * 100

        deleveraging = df[df["debt_change_pct"] <= -DELEVERAGING_DROP_PCT].copy()
        if deleveraging.empty:
            return pd.DataFrame()

        deleveraging["deleveraging_note"] = deleveraging["debt_change_pct"].apply(
            lambda v: f"Debt reduced by {abs(v):.1f}% YoY"
        )

        keep = ["company_id", "year", "total_debt_cr", "prev_debt",
                "debt_change_pct", "deleveraging_note"]
        return deleveraging[[c for c in keep if c in deleveraging.columns]].reset_index(drop=True)

    # ------------------------------------------------------------------
    # Pattern changes
    # ------------------------------------------------------------------

    def _detect_pattern_changes(self, health_df: pd.DataFrame) -> pd.DataFrame:
        """
        Detect companies whose capital allocation pattern changed YoY.
        """
        if health_df.empty:
            return pd.DataFrame()

        df = health_df.copy()
        df["year"] = pd.to_numeric(df["year"], errors="coerce")
        df = df.dropna(subset=["year"]).sort_values(["company_id", "year"])

        df["allocation_pattern"] = df.apply(self._classify_capital_allocation, axis=1)
        df["prev_pattern"] = df.groupby("company_id")["allocation_pattern"].shift(1)

        changes = df[
            df["prev_pattern"].notna() & (df["allocation_pattern"] != df["prev_pattern"])
        ].copy()

        if changes.empty:
            return pd.DataFrame()

        changes["change_description"] = (
            changes["prev_pattern"] + " → " + changes["allocation_pattern"]
        )

        keep = ["company_id", "year", "prev_pattern", "allocation_pattern", "change_description"]
        return changes[[c for c in keep if c in changes.columns]].reset_index(drop=True)

    # ------------------------------------------------------------------
    # Excel formatting
    # ------------------------------------------------------------------

    @staticmethod
    def _apply_excel_formatting(path: str) -> None:
        """Bold header, frozen pane, auto-width for every sheet."""
        try:
            wb = load_workbook(path)
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row == 0:
                    continue
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = header_fill
                ws.freeze_panes = "A2"
                for col_idx in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    max_len = max(
                        (len(str(cell.value)) for cell in ws[col_letter] if cell.value),
                        default=10,
                    )
                    ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
            wb.save(path)
        except Exception as exc:
            logger.warning("Could not apply Excel formatting: %s", exc)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def run(self) -> bool:
        """
        Execute the full Cash Flow Intelligence pipeline and write outputs.

        Returns
        -------
        True on success.
        """
        print("\n" + "=" * 70)
        print("CASH FLOW INTELLIGENCE ENGINE")
        print("=" * 70)

        cf_df = self._load_excel(self.cashflow_path, "cashflow.xlsx", header=1)
        pnl_df = self._load_excel(self.pnl_path, "profitandloss.xlsx", header=1)
        health_df = self._load_csv(self.health_scores_path, "company_health_scores.csv")

        # 1. CFO Quality
        cfo_quality_df = self._compute_cfo_quality(cf_df, pnl_df)
        logger.info("CFO quality: %d records.", len(cfo_quality_df))

        # 2. CapEx Intensity
        capex_df = self._compute_capex_intensity(health_df)
        logger.info("CapEx intensity: %d records.", len(capex_df))

        # 3. Capital Allocation Patterns
        if not health_df.empty:
            health_df["allocation_pattern"] = health_df.apply(
                self._classify_capital_allocation, axis=1
            )
            alloc_df = health_df[["company_id", "year", "allocation_pattern"]].copy()
        else:
            alloc_df = pd.DataFrame()

        # 4. Distress Detection
        distress_df = self._detect_distress(health_df)
        logger.info("Distress alerts: %d records.", len(distress_df))

        # 5. Deleveraging Detection
        deleveraging_df = self._detect_deleveraging(health_df)
        logger.info("Deleveraging events: %d records.", len(deleveraging_df))

        # 6. Pattern Changes
        pattern_changes_df = self._detect_pattern_changes(health_df)
        logger.info("Pattern changes: %d records.", len(pattern_changes_df))

        # Save pattern_changes.csv
        pattern_changes_df.to_csv(self.pattern_output_path, index=False)
        print(f"  Pattern changes  : {len(pattern_changes_df)}")
        print(f"  Saved: {self.pattern_output_path}")

        # Save distress_alerts.csv
        distress_df.to_csv(self.distress_output_path, index=False)
        print(f"  Distress alerts  : {len(distress_df)}")
        print(f"  Saved: {self.distress_output_path}")

        # Save cashflow_intelligence.xlsx
        sheets: dict = {}
        if not cfo_quality_df.empty:
            sheets["CFO Quality"] = cfo_quality_df
        if not capex_df.empty:
            sheets["CapEx Intensity"] = capex_df
        if not alloc_df.empty:
            sheets["Capital Allocation"] = alloc_df
        if not distress_df.empty:
            sheets["Distress Alerts"] = distress_df
        if not deleveraging_df.empty:
            sheets["Deleveraging"] = deleveraging_df
        if not pattern_changes_df.empty:
            sheets["Pattern Changes"] = pattern_changes_df

        if sheets:
            with pd.ExcelWriter(self.excel_output_path, engine="openpyxl") as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            self._apply_excel_formatting(self.excel_output_path)
            print(f"  Saved: {self.excel_output_path}")
        else:
            logger.warning("No data to write to cashflow_intelligence.xlsx.")

        print("Cash Flow Intelligence Engine complete.\n")
        return True


def _safe_float(val) -> Optional[float]:
    """Module-level safe float conversion."""
    if val is None:
        return None
    try:
        f = float(val)
        return None if pd.isna(f) else f
    except (ValueError, TypeError):
        return None


if __name__ == "__main__":
    engine = CashFlowIntelligenceEngine()
    engine.run()
