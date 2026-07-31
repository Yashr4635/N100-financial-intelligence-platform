"""
valuation.py

Valuation Engine for the N100 Financial Intelligence Platform.

Consumes the outputs of the Ratio Engine, Health Score Engine, and Sector
Analysis (Sprints 2-3), enriches them with sector mapping and market price
data (pe_ratio / pb_ratio from the 'market_cap' dataset), and produces a
single row of valuation metrics per company using each company's most
recent available fiscal year.

Inputs
------
- data/output/company_health_scores.csv   (ratios + health_score + rating)
- data/output/sector_analysis.csv         (avg_health_score per broad_sector)
- 'sectors' table (SQLite DB) with fallback to data/raw/sectors.xlsx
- 'market_cap' table (SQLite DB) with fallback to data/raw/market_cap.xlsx
- 'companies' table (SQLite DB), optional, used only for display names

Outputs
-------
- data/output/valuation_summary.xlsx   ('Valuation Summary' + 'Summary' sheets)
- data/output/valuation_flags.csv      (lean table: company, flags, score)

Design notes (read before changing thresholds)
------------------------------------------------
1. Unit of analysis is COMPANY, not company-year. Every other Sprint 1-3
   module operates on company-year rows (each company appears once per
   fiscal year in the ratio/health-score tables). A "valuation summary" is
   conventionally a current, as-of-today view, so this engine selects each
   company's latest available year before scoring. This is a deliberate
   deviation from the row-per-year convention used elsewhere in the
   pipeline, not an oversight.
2. P/E and P/B are pulled from the 'market_cap' dataset (pe_ratio /
   pb_ratio), matched on (company_id, year) to the company's latest ratio
   year. financial_ratios_calculated.csv / company_health_scores.csv do not
   contain price data, so this is unavoidable if P/E and P/B are to mean
   anything (a P/E computed from an EPS in one year and a price from an
   unrelated year is not a valid ratio).
3. "Undervalued" / "Overvalued" are price-relative judgements. They cannot
   honestly be assigned to a company with no market price. Rather than
   forcing every row into Undervalued/Overvalued regardless of data
   availability, this engine defaults such rows to "Fairly Valued" (the
   least-wrong neutral label) and exposes a separate `price_data_available`
   column so any downstream consumer can filter out or flag calls that are
   not price-based. Treat `valuation_basis` == "Quality-Proxy (No Market
   Price)" rows as informational only, not investment signals.

Dependencies: pandas, openpyxl. sqlite3 is optional (falls back to raw
Excel files if the database is missing or a table doesn't exist).
"""

from __future__ import annotations

import logging
import os
import sqlite3
from typing import Dict, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)
if not logger.handlers:
    logging.basicConfig(level=logging.INFO)


class ValuationEngine:
    """
    Computes a per-company valuation summary combining fundamental quality
    (health score, financial quality score, FCF yield) with market-relative
    valuation (P/E and P/B vs. sector averages), where price data exists.

    See module docstring for the "unit of analysis" and "no market price"
    design decisions before changing any scoring weights or thresholds.
    """

    HEALTH_SCORES_PATH = os.path.join("data", "output", "company_health_scores.csv")
    SECTOR_ANALYSIS_PATH = os.path.join("data", "output", "sector_analysis.csv")
    SECTORS_XLSX_FALLBACK = os.path.join("data", "raw", "sectors.xlsx")
    MARKET_CAP_XLSX_FALLBACK = os.path.join("data", "raw", "market_cap.xlsx")
    COMPANIES_XLSX_FALLBACK = os.path.join("data", "raw", "companies.xlsx")
    OUTPUT_DIR = os.path.join("data", "output")

    # Composite Valuation Score weights. Any component that is unavailable
    # for a given row (missing data) is dropped and the remaining weights
    # are renormalized to sum to 1.0 for that row, rather than silently
    # treating a missing metric as zero.
    WEIGHTS = {
        "health": 0.35,
        "quality": 0.15,
        "fcf_yield": 0.20,
        "relative_value": 0.30,
    }

    # Threshold (in %) beyond which a company is considered meaningfully
    # cheaper/pricier than its sector on a blended P/E + P/B basis.
    RELATIVE_VALUE_BAND_PCT = 15.0

    def __init__(self, database_path: Optional[str] = None, output_dir: Optional[str] = None):
        try:
            from src.utils.config import DATABASE_PATH  # local import: avoid hard dependency
            self.database_path = database_path or str(DATABASE_PATH)
        except Exception:
            self.database_path = database_path

        self.output_dir = output_dir or self.OUTPUT_DIR
        os.makedirs(self.output_dir, exist_ok=True)

        self.excel_output_path = os.path.join(self.output_dir, "valuation_summary.xlsx")
        self.flags_output_path = os.path.join(self.output_dir, "valuation_flags.csv")

        self.df: pd.DataFrame = pd.DataFrame()
        self.summary: Dict[str, object] = {}

    # ------------------------------------------------------------------ #
    # Defensive loading helpers
    # ------------------------------------------------------------------ #

    @staticmethod
    def _load_csv(path: str, label: str) -> pd.DataFrame:
        """Load a CSV safely, returning an empty DataFrame if missing/unreadable."""
        if not os.path.exists(path):
            logger.warning("%s not found at '%s'. Using an empty DataFrame.", label, path)
            return pd.DataFrame()
        try:
            df = pd.read_csv(path)
            logger.info("Loaded %s: %d rows, %d columns.", label, len(df), len(df.columns))
            return df
        except Exception as exc:  # noqa: BLE001 - defensive load, must not crash the pipeline
            logger.warning("Failed to read %s from '%s': %s. Using an empty DataFrame.", label, path, exc)
            return pd.DataFrame()

    # Files that carry a metadata title row above the real header, matching
    # ExcelLoader.CORE_DATASETS in src/etl/loader.py. Only relevant to the
    # raw-Excel fallback path; the database path already has clean headers.
    _CORE_DATASET_STEMS = {
        "companies", "profitandloss", "balancesheet", "cashflow",
        "analysis", "documents", "prosandcons",
    }

    def _load_table(self, table_name: str, xlsx_fallback: str, label: str) -> pd.DataFrame:
        """
        Load a dataset by name, preferring the SQLite database (consistent
        with how SectorAnalysis / PeerComparisonEngine source sector data)
        and falling back to the raw Excel file if the DB or table is
        unavailable. Never raises; returns an empty DataFrame on failure.
        """
        if self.database_path and os.path.exists(self.database_path):
            try:
                conn = sqlite3.connect(self.database_path)
                try:
                    df = pd.read_sql(f"SELECT * FROM {table_name}", conn)
                    logger.info("Loaded %s from database table '%s': %d rows.", label, table_name, len(df))
                    return df
                finally:
                    conn.close()
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "Could not read table '%s' from database (%s). Falling back to '%s'.",
                    table_name, exc, xlsx_fallback,
                )

        if os.path.exists(xlsx_fallback):
            try:
                stem = os.path.splitext(os.path.basename(xlsx_fallback))[0]
                header_row = 1 if stem in self._CORE_DATASET_STEMS else 0
                df = pd.read_excel(xlsx_fallback, header=header_row)
                df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_").str.replace("-", "_")
                if "company_id" in df.columns:
                    df["company_id"] = df["company_id"].astype(str).str.strip().str.upper()
                logger.info("Loaded %s from fallback file '%s': %d rows.", label, xlsx_fallback, len(df))
                return df
            except Exception as exc:  # noqa: BLE001
                logger.warning("Failed to read fallback '%s': %s.", xlsx_fallback, exc)

        logger.warning("%s unavailable (no database table and no fallback file). Continuing without it.", label)
        return pd.DataFrame()

    # ------------------------------------------------------------------ #
    # Pipeline steps
    # ------------------------------------------------------------------ #

    def _select_latest_year_per_company(self, health_df: pd.DataFrame) -> pd.DataFrame:
        """Keep only each company's most recent available fiscal year row."""
        if health_df.empty or "company_id" not in health_df.columns or "year" not in health_df.columns:
            return health_df

        health_df = health_df.copy()
        health_df["year"] = pd.to_numeric(health_df["year"], errors="coerce")
        latest_idx = health_df.groupby("company_id")["year"].idxmax().dropna()
        return health_df.loc[latest_idx].reset_index(drop=True)

    def _attach_sector(self, df: pd.DataFrame) -> pd.DataFrame:
        """Merge in broad_sector per company_id, and sector-level avg health score."""
        sectors_df = self._load_table("sectors", self.SECTORS_XLSX_FALLBACK, "sector mapping")
        sector_analysis_df = self._load_csv(self.SECTOR_ANALYSIS_PATH, "sector_analysis.csv")

        if not sectors_df.empty and "company_id" in sectors_df.columns and "broad_sector" in sectors_df.columns:
            df = df.merge(sectors_df[["company_id", "broad_sector"]], on="company_id", how="left")
        else:
            logger.warning("Sector mapping unavailable; 'broad_sector' will be missing for all companies.")
            df["broad_sector"] = pd.NA

        if not sector_analysis_df.empty and "broad_sector" in sector_analysis_df.columns:
            sector_avg = sector_analysis_df[["broad_sector", "avg_health_score"]].rename(
                columns={"avg_health_score": "sector_avg_health_score"}
            )
            df = df.merge(sector_avg, on="broad_sector", how="left")
        else:
            logger.warning("sector_analysis.csv unavailable/malformed; 'sector_avg_health_score' will be missing.")
            df["sector_avg_health_score"] = pd.NA

        return df

    def _attach_market_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Merge in pe_ratio, pb_ratio, market_cap_crore matched on
        (company_id, year). Also computes each sector's average P/E and
        average P/B (across all years present in market_cap data) so
        relative valuation has something to compare against.
        """
        market_df = self._load_table("market_cap", self.MARKET_CAP_XLSX_FALLBACK, "market cap / price data")

        if market_df.empty or "company_id" not in market_df.columns:
            logger.warning("Market price data unavailable. P/E, P/B, and FCF Yield will be 'Unavailable'.")
            df["pe_ratio"] = pd.NA
            df["pb_ratio"] = pd.NA
            df["market_cap_crore"] = pd.NA
            df["sector_avg_pe"] = pd.NA
            df["sector_avg_pb"] = pd.NA
            return df

        market_df = market_df.copy()
        market_df["year"] = pd.to_numeric(market_df["year"], errors="coerce")

        # Company-level, year-matched P/E, P/B, market cap.
        df = df.merge(
            market_df[["company_id", "year", "pe_ratio", "pb_ratio", "market_cap_crore"]],
            on=["company_id", "year"],
            how="left",
        )

        # Sector-level average P/E and P/B, for relative valuation. Needs
        # the same sector mapping used for health scores.
        sectors_df = self._load_table("sectors", self.SECTORS_XLSX_FALLBACK, "sector mapping")
        if not sectors_df.empty and "broad_sector" in sectors_df.columns:
            market_with_sector = market_df.merge(
                sectors_df[["company_id", "broad_sector"]], on="company_id", how="left"
            )
            sector_pe_pb = (
                market_with_sector.groupby("broad_sector")[["pe_ratio", "pb_ratio"]]
                .mean()
                .rename(columns={"pe_ratio": "sector_avg_pe", "pb_ratio": "sector_avg_pb"})
                .reset_index()
            )
            df = df.merge(sector_pe_pb, on="broad_sector", how="left")
        else:
            df["sector_avg_pe"] = pd.NA
            df["sector_avg_pb"] = pd.NA

        return df

    def _attach_company_names(self, df: pd.DataFrame) -> pd.DataFrame:
        """Best-effort merge of a readable company name. Never blocks the pipeline."""
        companies_df = self._load_table("companies", self.COMPANIES_XLSX_FALLBACK, "company names")
        if companies_df.empty:
            df["company_name"] = df["company_id"]
            return df

        name_col = next((c for c in ("company_name", "name") if c in companies_df.columns), None)
        id_col = next((c for c in ("id", "company_id") if c in companies_df.columns), None)

        if not name_col or not id_col:
            df["company_name"] = df["company_id"]
            return df

        lookup = companies_df[[id_col, name_col]].rename(columns={id_col: "company_id", name_col: "company_name"})
        lookup["company_id"] = lookup["company_id"].astype(str).str.strip().str.upper()
        df = df.merge(lookup, on="company_id", how="left")
        df["company_name"] = df["company_name"].fillna(df["company_id"])
        return df

    def _compute_fcf_yield(self, df: pd.DataFrame) -> pd.DataFrame:
        """FCF Yield (%) = Free Cash Flow / Market Cap * 100. NaN if either is missing/zero."""
        has_inputs = df["free_cash_flow_cr"].notna() & df["market_cap_crore"].notna() & (df["market_cap_crore"] > 0)
        df["fcf_yield_pct"] = pd.NA
        df.loc[has_inputs, "fcf_yield_pct"] = (
            df.loc[has_inputs, "free_cash_flow_cr"] / df.loc[has_inputs, "market_cap_crore"] * 100
        ).round(2)
        return df

    def _compute_relative_valuation(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Blended discount/premium (%) of company P/E and P/B vs. sector
        averages. Positive = cheaper than sector, negative = pricier.
        Uses whichever of P/E / P/B is available; NaN if neither is.
        """
        def _discount(company_val, sector_val):
            if pd.isna(company_val) or pd.isna(sector_val) or sector_val == 0:
                return pd.NA
            return (sector_val - company_val) / sector_val * 100

        pe_discount = df.apply(lambda r: _discount(r["pe_ratio"], r["sector_avg_pe"]), axis=1)
        pb_discount = df.apply(lambda r: _discount(r["pb_ratio"], r["sector_avg_pb"]), axis=1)

        blended = pd.concat([pe_discount, pb_discount], axis=1)
        blended.columns = ["pe_discount", "pb_discount"]
        df["relative_discount_pct"] = blended.mean(axis=1, skipna=True).round(2)

        def _label(discount):
            if pd.isna(discount):
                return "Not Determinable - No Market Price"
            if discount > self.RELATIVE_VALUE_BAND_PCT:
                return "Below Sector Average (Cheaper)"
            if discount < -self.RELATIVE_VALUE_BAND_PCT:
                return "Above Sector Average (Expensive)"
            return "In Line With Sector"

        df["relative_valuation"] = df["relative_discount_pct"].apply(_label)
        return df

    def _compute_valuation_score(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Weighted composite score (0-100) blending fundamental quality and
        market-relative cheapness. Missing components are dropped and the
        remaining weights renormalized per row (see WEIGHTS / module docstring).
        """
        def _quality_score(row):
            fq = row.get("financial_quality_score")
            return None if pd.isna(fq) else max(0.0, min(100.0, float(fq) * 20.0))

        def _fcf_score(row):
            y = row.get("fcf_yield_pct")
            if pd.isna(y):
                return None
            return max(0.0, min(100.0, float(y) / 20.0 * 100.0))

        def _relative_value_score(row):
            d = row.get("relative_discount_pct")
            if pd.isna(d):
                return None
            return 50.0 + max(-50.0, min(50.0, float(d)))

        def _score_row(row):
            components = {
                "health": None if pd.isna(row.get("health_score")) else float(row["health_score"]),
                "quality": _quality_score(row),
                "fcf_yield": _fcf_score(row),
                "relative_value": _relative_value_score(row),
            }
            available = {k: v for k, v in components.items() if v is not None}
            if not available:
                return pd.NA

            weight_sum = sum(self.WEIGHTS[k] for k in available)
            weighted = sum(self.WEIGHTS[k] * v for k, v in available.items())
            return round(weighted / weight_sum, 2)

        df["valuation_score"] = df.apply(_score_row, axis=1)
        return df

    def _assign_valuation_flag(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Assigns Undervalued / Fairly Valued / Overvalued.

        Rows without market price data cannot be honestly called
        Undervalued or Overvalued (there is no price to compare against),
        so they default to Fairly Valued with `valuation_basis` marked as
        a quality-only proxy. Treat those rows as informational, not as
        buy/sell signals.
        """
        df["price_data_available"] = df["pe_ratio"].notna() | df["pb_ratio"].notna()

        def _flag(row):
            score = row["valuation_score"]
            score = None if pd.isna(score) else float(score)

            if not row["price_data_available"]:
                return "Fairly Valued"
            if score is None:
                return "Fairly Valued"

            if row["relative_valuation"] == "Below Sector Average (Cheaper)" and score >= 50:
                return "Undervalued"
            if row["relative_valuation"] == "Above Sector Average (Expensive)" or score < 35:
                return "Overvalued"
            return "Fairly Valued"

        df["valuation_flag"] = df.apply(_flag, axis=1)
        df["valuation_basis"] = df["price_data_available"].map(
            {True: "Price-Based", False: "Quality-Proxy (No Market Price)"}
        )
        return df

    @staticmethod
    def _format_unavailable(df: pd.DataFrame) -> pd.DataFrame:
        """Replace NaN in price-dependent display columns with the literal string 'Unavailable'."""
        display_cols = ["pe_ratio", "pb_ratio", "fcf_yield_pct", "sector_avg_pe", "sector_avg_pb"]
        for col in display_cols:
            if col in df.columns:
                df[col] = df[col].apply(lambda v: "Unavailable" if pd.isna(v) else v)
        return df

    # ------------------------------------------------------------------ #
    # Output
    # ------------------------------------------------------------------ #

    def _build_summary_stats(self, df: pd.DataFrame) -> Dict[str, object]:
        total = len(df)
        priced = int(df["price_data_available"].sum()) if "price_data_available" in df.columns else 0

        summary: Dict[str, object] = {
            "Total Companies": total,
            "Companies With Market Price Data": priced,
            "Companies Without Market Price Data": total - priced,
        }

        if "valuation_flag" in df.columns:
            counts = df["valuation_flag"].value_counts()
            for label in ("Undervalued", "Fairly Valued", "Overvalued"):
                summary[f"{label} (count)"] = int(counts.get(label, 0))

        if "valuation_score" in df.columns:
            scores = pd.to_numeric(df["valuation_score"], errors="coerce").dropna()
            summary["Average Valuation Score"] = round(float(scores.mean()), 2) if not scores.empty else "N/A"

        self.summary = summary
        return summary

    def _write_excel(self, df: pd.DataFrame) -> str:
        summary_df = pd.DataFrame(
            {"Metric": list(self.summary.keys()), "Value": pd.Series(list(self.summary.values()), dtype=object)}
        )

        with pd.ExcelWriter(self.excel_output_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            if not df.empty:
                df.to_excel(writer, sheet_name="Valuation Summary", index=False)
            else:
                pd.DataFrame({"Message": ["No valuation data available."]}).to_excel(
                    writer, sheet_name="Valuation Summary", index=False
                )

        self._apply_formatting(self.excel_output_path)
        logger.info("Saved valuation workbook to '%s'.", self.excel_output_path)
        return self.excel_output_path

    @staticmethod
    def _apply_formatting(path: str) -> None:
        """Bold header row, frozen header, auto-width columns (matches ReportingEngine style)."""
        wb = load_workbook(path)
        header_font = Font(bold=True)

        for sheet_name in wb.sheetnames:
            ws: Worksheet = wb[sheet_name]
            if ws.max_row == 0 or ws.max_column == 0:
                continue

            for cell in ws[1]:
                cell.font = header_font
            ws.freeze_panes = "A2"

            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in ws[col_letter]:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 60)

        wb.save(path)

    def _write_flags_csv(self, df: pd.DataFrame) -> str:
        flag_cols = [
            "company_id", "company_name", "broad_sector", "year",
            "health_score", "sector_avg_health_score",
            "pe_ratio", "pb_ratio", "relative_valuation",
            "valuation_score", "valuation_flag", "valuation_basis", "price_data_available",
        ]
        available_cols = [c for c in flag_cols if c in df.columns]
        df[available_cols].to_csv(self.flags_output_path, index=False)
        logger.info("Saved valuation flags to '%s'.", self.flags_output_path)
        return self.flags_output_path

    # ------------------------------------------------------------------ #
    # Public API
    # ------------------------------------------------------------------ #

    def run(self) -> pd.DataFrame:
        """
        Executes the full valuation pipeline and writes both output files.
        Returns the enriched per-company valuation DataFrame (with real NaN
        values, not the 'Unavailable' strings used in the Excel/CSV output,
        so callers can still do numeric filtering on the returned frame).
        """
        print("\n" + "=" * 70)
        print("VALUATION ENGINE")
        print("=" * 70)

        health_df = self._load_csv(self.HEALTH_SCORES_PATH, "company_health_scores.csv")
        if health_df.empty:
            logger.error("company_health_scores.csv is required and unavailable. Aborting valuation run.")
            self.df = pd.DataFrame()
            self._build_summary_stats(self.df)
            self._write_excel(self.df)
            self._write_flags_csv(self.df)
            return self.df

        df = self._select_latest_year_per_company(health_df)
        df = self._attach_sector(df)
        df = self._attach_market_data(df)
        df = self._attach_company_names(df)
        df = self._compute_fcf_yield(df)
        df = self._compute_relative_valuation(df)
        df = self._compute_valuation_score(df)
        df = self._assign_valuation_flag(df)

        numeric_df = df.copy()  # retained internally with real NaNs, pre-formatting

        display_df = self._format_unavailable(df.copy())
        display_df = display_df.sort_values(by="valuation_score", ascending=False, na_position="last")

        self._build_summary_stats(numeric_df)
        self._write_excel(display_df)
        self._write_flags_csv(display_df)

        print(f"Companies scored: {len(display_df)}")
        print(f"With market price data: {self.summary.get('Companies With Market Price Data', 'N/A')}")
        print(f"Saved: {self.excel_output_path}")
        print(f"Saved: {self.flags_output_path}")
        print("\nValuation complete.")

        self.df = numeric_df
        return self.df


if __name__ == "__main__":
    # Production entry point: runs the valuation engine against the
    # existing project outputs (data/output/company_health_scores.csv,
    # data/output/sector_analysis.csv) and the existing database / raw
    # Excel files for sector mapping and market price data.
    #
    # This block intentionally does NOT generate or write any sample data.
    # Earlier versions of this file wrote sample rows (ALPHA/BETA/GAMMA) to
    # data/output/company_health_scores.csv and data/output/sector_analysis.csv
    # for smoke testing, which silently overwrote the real ~92-company
    # analytics outputs on every run. That code path has been removed.
    engine = ValuationEngine()
    engine.run()