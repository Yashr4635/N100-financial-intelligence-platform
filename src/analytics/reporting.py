"""
reporting.py

Reporting Engine for the N100 Financial Intelligence Platform.

Consumes the outputs of the Investment Screener and Peer Comparison Engine
and produces two consolidated reports:

  - data/output/analytics_summary.xlsx
      * 'Summary'              high-level portfolio statistics
      * 'Investment Screener'  full screener output
      * 'Peer Comparison'      full peer comparison output
  - data/output/executive_summary.csv
      Flat metric/value view of the Summary sheet.

Dependencies: pandas, openpyxl only.
"""

from __future__ import annotations

import logging
import os
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)
if not logger.handlers:
    logging.basicConfig(level=logging.INFO)


class ReportingEngine:
    """
    Builds consolidated analytics reports from the Investment Screener and
    Peer Comparison Engine outputs.

    Inputs
    ------
    - data/output/investment_screener.csv
    - data/output/peer_comparison.csv

    Both inputs are read defensively: a missing file or read error logs a
    warning and falls back to an empty DataFrame rather than raising, so
    the reporting pipeline can still produce partial output.

    Outputs
    -------
    - data/output/analytics_summary.xlsx
    - data/output/executive_summary.csv
    """

    SCREENER_INPUT_PATH = os.path.join("data", "output", "investment_screener.csv")
    PEER_INPUT_PATH = os.path.join("data", "output", "peer_comparison.csv")
    OUTPUT_DIR = os.path.join("data", "output")

    # Candidate source-column names for summary statistics, checked in
    # priority order so the engine tolerates minor upstream naming changes.
    SECTOR_COL_CANDIDATES = ["sector", "broad_sector", "Sector", "Broad_Sector", "sector_name"]
    HEALTH_SCORE_CANDIDATES = ["health_score"]
    ROE_CANDIDATES = ["roe", "return_on_equity_pct", "return_on_equity"]
    NET_MARGIN_CANDIDATES = ["net_profit_margin", "net_profit_margin_pct"]
    DEBT_TO_EQUITY_CANDIDATES = ["debt_to_equity"]

    def __init__(
        self,
        screener_path: Optional[str] = None,
        peer_path: Optional[str] = None,
        output_dir: Optional[str] = None,
    ):
        """
        Load the Investment Screener and Peer Comparison outputs and
        prepare the output directory.

        Parameters
        ----------
        screener_path : optional override for the investment screener CSV path.
        peer_path : optional override for the peer comparison CSV path.
        output_dir : optional override for the report output directory.
        """
        self.screener_path = screener_path or self.SCREENER_INPUT_PATH
        self.peer_path = peer_path or self.PEER_INPUT_PATH
        self.output_dir = output_dir or self.OUTPUT_DIR

        os.makedirs(self.output_dir, exist_ok=True)

        self.excel_output_path = os.path.join(self.output_dir, "analytics_summary.xlsx")
        self.exec_summary_output_path = os.path.join(self.output_dir, "executive_summary.csv")

        self.screener_df: pd.DataFrame = self._load_csv(self.screener_path, "Investment Screener")
        self.peer_df: pd.DataFrame = self._load_csv(self.peer_path, "Peer Comparison")

        self.summary: Dict[str, object] = {}

    # ------------------------------------------------------------------ #
    # Loading helpers
    # ------------------------------------------------------------------ #

    @staticmethod
    def _load_csv(path: str, label: str) -> pd.DataFrame:
        """Load a CSV safely, returning an empty DataFrame if missing/unreadable."""
        if not os.path.exists(path):
            logger.warning("%s input not found at '%s'. Using an empty DataFrame.", label, path)
            return pd.DataFrame()
        try:
            df = pd.read_csv(path)
            logger.info(
                "Loaded %s: %d rows, %d columns from '%s'.", label, len(df), len(df.columns), path
            )
            return df
        except Exception as exc:  # noqa: BLE001 - defensive load, must not crash the pipeline
            logger.warning("Failed to read %s from '%s': %s. Using an empty DataFrame.", label, path, exc)
            return pd.DataFrame()

    @staticmethod
    def _resolve_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
        """Return the first candidate column name present in df, else None."""
        for candidate in candidates:
            if candidate in df.columns:
                return candidate
        return None

    # ------------------------------------------------------------------ #
    # Summary generation
    # ------------------------------------------------------------------ #

    def generate_summary(self) -> Dict[str, object]:
        """
        Compute high-level portfolio statistics.

        'Total Companies', 'Number of Sectors', and the metric averages are
        derived from the Peer Comparison output (the full company universe).
        'Screened Companies' is the row count of the Investment Screener
        output. Any metric whose source column cannot be found is reported
        as 'N/A' rather than raising an error.

        Returns
        -------
        dict mapping metric label -> value, in display order.
        """
        summary: Dict[str, object] = {}
        peer_df = self.peer_df
        screener_df = self.screener_df

        # Total Companies: prefer the full peer-comparison universe; fall
        # back to the screener count if peer data is unavailable.
        if not peer_df.empty:
            summary["Total Companies"] = int(len(peer_df))
        elif not screener_df.empty:
            summary["Total Companies"] = int(len(screener_df))
        else:
            summary["Total Companies"] = "N/A"

        summary["Screened Companies"] = int(len(screener_df)) if not screener_df.empty else "N/A"

        sector_col = self._resolve_column(peer_df, self.SECTOR_COL_CANDIDATES)
        if sector_col:
            summary["Number of Sectors"] = int(peer_df[sector_col].nunique(dropna=True))
        else:
            summary["Number of Sectors"] = "N/A"
            logger.warning("No sector column found in Peer Comparison data; skipping 'Number of Sectors'.")

        self._add_stat_metrics(
            summary,
            peer_df,
            column_candidates=self.HEALTH_SCORE_CANDIDATES,
            metric_labels=("Average Health Score", "Highest Health Score", "Lowest Health Score"),
            stats=("mean", "max", "min"),
        )
        self._add_stat_metrics(
            summary, peer_df, column_candidates=self.ROE_CANDIDATES, metric_labels=("Average ROE",), stats=("mean",)
        )
        self._add_stat_metrics(
            summary,
            peer_df,
            column_candidates=self.NET_MARGIN_CANDIDATES,
            metric_labels=("Average Net Profit Margin",),
            stats=("mean",),
        )
        self._add_stat_metrics(
            summary,
            peer_df,
            column_candidates=self.DEBT_TO_EQUITY_CANDIDATES,
            metric_labels=("Average Debt to Equity",),
            stats=("mean",),
        )

        self.summary = summary
        logger.info("Summary statistics computed: %s", summary)
        return summary

    def _add_stat_metrics(
        self,
        summary: Dict[str, object],
        df: pd.DataFrame,
        column_candidates: List[str],
        metric_labels: tuple,
        stats: tuple,
    ) -> None:
        """
        Resolve a source column from candidates and add one or more
        statistics (mean/max/min) to `summary` under the given labels.
        Missing columns or all-NaN data are recorded as 'N/A'.
        """
        source_col = self._resolve_column(df, column_candidates)

        if not source_col:
            for label in metric_labels:
                summary[label] = "N/A"
            logger.warning(
                "No column found for %s among candidates %s.", metric_labels, column_candidates
            )
            return

        series = pd.to_numeric(df[source_col], errors="coerce").dropna()

        for label, stat in zip(metric_labels, stats):
            if series.empty:
                summary[label] = "N/A"
                continue
            value = getattr(series, stat)()
            summary[label] = round(float(value), 2)

    # ------------------------------------------------------------------ #
    # Excel generation
    # ------------------------------------------------------------------ #

    def generate_excel(self) -> str:
        """
        Build 'analytics_summary.xlsx' with three sheets:
          - Summary               metric/value pairs
          - Investment Screener   full screener output
          - Peer Comparison       full peer comparison output

        Applies basic formatting to every sheet: bold header row, frozen
        first row, and auto-adjusted column widths.

        Returns
        -------
        The path to the generated workbook.
        """
        if not self.summary:
            self.generate_summary()

        summary_df = self._summary_dict_to_frame()

        with pd.ExcelWriter(self.excel_output_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            self._write_sheet_or_placeholder(
                writer, self.screener_df, "Investment Screener", "No investment screener data available."
            )
            self._write_sheet_or_placeholder(
                writer, self.peer_df, "Peer Comparison", "No peer comparison data available."
            )

        self._apply_formatting(self.excel_output_path)
        logger.info("Saved analytics summary workbook to '%s'.", self.excel_output_path)
        return self.excel_output_path

    @staticmethod
    def _write_sheet_or_placeholder(
        writer: "pd.ExcelWriter", df: pd.DataFrame, sheet_name: str, placeholder_message: str
    ) -> None:
        """Write `df` to `sheet_name`, or a one-row placeholder message if empty."""
        if not df.empty:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            pd.DataFrame({"Message": [placeholder_message]}).to_excel(
                writer, sheet_name=sheet_name, index=False
            )

    def _summary_dict_to_frame(self) -> pd.DataFrame:
        """
        Convert the summary dict into a two-column 'Metric'/'Value' DataFrame.

        The 'Value' column is built with dtype=object so integer metrics
        (e.g. 'Total Companies') are not silently upcast to float just
        because other metrics in the same column are floats or 'N/A'.
        """
        metrics = list(self.summary.keys())
        values = pd.Series(list(self.summary.values()), dtype=object)
        return pd.DataFrame({"Metric": metrics, "Value": values})

    @staticmethod
    def _apply_formatting(path: str) -> None:
        """
        Apply consistent basic formatting to every sheet in the workbook:
        bold header row, frozen header row, and auto-adjusted column widths
        (capped to keep very long text fields readable).
        """
        wb = load_workbook(path)
        header_font = Font(bold=True)

        for sheet_name in wb.sheetnames:
            ws: Worksheet = wb[sheet_name]
            if ws.max_row == 0 or ws.max_column == 0:
                continue

            # Bold header row.
            for cell in ws[1]:
                cell.font = header_font

            # Freeze the header row so it stays visible while scrolling.
            ws.freeze_panes = "A2"

            # Auto-adjust column widths based on the longest value in each column.
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in ws[col_letter]:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 60)

        wb.save(path)

    # ------------------------------------------------------------------ #
    # Executive summary CSV
    # ------------------------------------------------------------------ #

    def generate_executive_summary_csv(self) -> str:
        """
        Write the summary statistics as a flat metric/value CSV.

        Returns
        -------
        The path to the generated CSV.
        """
        if not self.summary:
            self.generate_summary()

        summary_df = self._summary_dict_to_frame()
        summary_df.to_csv(self.exec_summary_output_path, index=False)
        logger.info("Saved executive summary to '%s'.", self.exec_summary_output_path)
        return self.exec_summary_output_path

    # ------------------------------------------------------------------ #
    # Public API
    # ------------------------------------------------------------------ #

    def run(self) -> bool:
        """
        Execute the full reporting pipeline:
          1. Compute summary statistics.
          2. Generate the Excel workbook (Summary / Screener / Peer Comparison).
          3. Generate the executive summary CSV.

        Returns
        -------
        True on successful completion.
        """
        print("\n" + "=" * 70)
        print("REPORTING ENGINE")
        print("=" * 70)

        self.generate_summary()
        excel_path = self.generate_excel()
        csv_path = self.generate_executive_summary_csv()

        print(f"Saved: {excel_path}")
        print(f"Saved: {csv_path}")
        print("\nReporting complete.")

        return True


if __name__ == "__main__":
    # Manual smoke test / usage example using sample data written to
    # data/output/, mirroring the real screener + peer comparison outputs.
    os.makedirs("data/output", exist_ok=True)

    sample_screener = pd.DataFrame(
        {
            "company_id": [1, 2, 3],
            "company": ["Alpha", "Beta", "Gamma"],
            "sector": ["Tech", "Tech", "Finance"],
            "health_score": [80, 90, 70],
        }
    )
    sample_screener.to_csv("data/output/investment_screener.csv", index=False)

    sample_peer = pd.DataFrame(
        {
            "company_id": [1, 2, 3, 4, 5],
            "sector": ["Tech", "Tech", "Tech", "Finance", "Finance"],
            "return_on_equity_pct": [15.0, 22.5, 9.0, 11.0, 13.0],
            "net_profit_margin_pct": [12.0, 18.0, 20.0, 15.0, 10.0],
            "debt_to_equity": [0.5, 0.3, 1.2, 0.9, 0.7],
            "health_score": [80, 90, 70, 65, 75],
        }
    )
    sample_peer.to_csv("data/output/peer_comparison.csv", index=False)

    engine = ReportingEngine()
    success = engine.run()
    print(f"\nrun() returned: {success}")